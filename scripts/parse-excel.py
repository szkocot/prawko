#!/usr/bin/env python3
"""
Parse Polish driving exam question Excel file and generate JSON data files.

Generates:
  - src/data/meta.json    — category metadata with counts and exam rules
  - src/data/{cat}.json   — per-category question banks

Usage:
  python3 scripts/parse-excel.py [--excel PATH] [--media-dir PATH] [--out-dir PATH]
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CATEGORIES = ["A", "A1", "A2", "AM", "B", "B1", "C", "C1", "D", "D1", "PT", "T"]

EXAM_RULES = {
    "totalQuestions": 32,
    "basicQuestions": 20,
    "specialistQuestions": 12,
    "maxPoints": 74,
    "passThreshold": 68,
    "totalTimeSeconds": 1500,
    "basicTimeSeconds": 20,
    "specialistTimeSeconds": 50,
    "basicPoints": [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 2, 2, 2, 2, 2, 2, 1, 1, 1, 1],
    "specialistPoints": [3, 3, 3, 3, 3, 3, 2, 2, 2, 2, 1, 1],
}

# Column indices (0-based) in the Excel sheet
COL_LP = 0          # L.p.
COL_NUM = 1         # Numer pytania
COL_Q = 2           # Pytanie
COL_A = 3           # Odpowiedź A
COL_B = 4           # Odpowiedź B
COL_C = 5           # Odpowiedź C
COL_CORRECT = 6     # Poprawna odp
COL_MEDIA = 7       # Media
COL_STRUCTURE = 8   # Zakres struktury (PODSTAWOWY / SPECJALISTYCZNY)
COL_CATEGORIES = 9  # Kategorie (comma-separated)

# ---------------------------------------------------------------------------
# Media helpers
# ---------------------------------------------------------------------------
MEDIA_EXT_MAP = {
    ".wmv": (".mp4", "video"),
    ".jpg": (".webp", "image"),
    ".jpeg": (".webp", "image"),
}


def resolve_media(raw_filename: str | None, media_dir: Path | None):
    """Return (target_filename, mediaType) or (None, None)."""
    if not raw_filename or not str(raw_filename).strip():
        return None, None

    raw_filename = str(raw_filename).strip()
    src_ext = os.path.splitext(raw_filename)[1].lower()
    mapping = MEDIA_EXT_MAP.get(src_ext)

    if mapping is None:
        # Unknown extension — keep as-is but warn
        print(f"  WARNING: unknown media extension '{src_ext}' for '{raw_filename}'")
        return raw_filename, "unknown"

    target_ext, media_type = mapping
    target_name = os.path.splitext(raw_filename)[0] + target_ext

    # Check if the source file exists
    if media_dir is not None:
        # Try exact match first, then case-insensitive
        src_path = media_dir / raw_filename
        if not src_path.exists():
            # Try case-insensitive lookup
            found = False
            for f in media_dir.iterdir():
                if f.name.lower() == raw_filename.lower():
                    found = True
                    break
            if not found:
                return None, None

    return target_name, media_type


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel",
        default="Pytania_egzaminacyjne_na_kierowcę_122025.xlsx",
        help="Path to the Excel question bank",
    )
    parser.add_argument(
        "--media-dir",
        default="Pytania egzaminacyjne na prawo jazdy 2025",
        help="Directory with source media files",
    )
    parser.add_argument(
        "--out-dir",
        default="src/data",
        help="Output directory for JSON files",
    )
    args = parser.parse_args()

    # Resolve paths relative to project root (parent of scripts/)
    project_root = Path(__file__).resolve().parent.parent
    excel_path = Path(args.excel) if os.path.isabs(args.excel) else project_root / args.excel
    media_dir = Path(args.media_dir) if os.path.isabs(args.media_dir) else project_root / args.media_dir
    out_dir = Path(args.out_dir) if os.path.isabs(args.out_dir) else project_root / args.out_dir

    if not excel_path.exists():
        sys.exit(f"Excel file not found: {excel_path}")
    if not media_dir.exists():
        print(f"WARNING: media directory not found: {media_dir} — skipping media checks")
        media_dir_for_check = None
    else:
        media_dir_for_check = media_dir

    out_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    print(f"Loading {excel_path.name} ...")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    data_rows = rows[1:]
    print(f"  {len(data_rows)} questions found (header: {len(header)} cols)")

    # Build per-category question lists
    cat_questions: dict[str, list] = {cat: [] for cat in CATEGORIES}
    missing_media_count = 0

    for row in data_rows:
        qnum = str(row[COL_NUM]).strip() if row[COL_NUM] is not None else ""
        question_text = str(row[COL_Q]).strip() if row[COL_Q] else ""
        correct = str(row[COL_CORRECT]).strip() if row[COL_CORRECT] else ""
        structure = str(row[COL_STRUCTURE]).strip() if row[COL_STRUCTURE] else ""
        raw_cats = str(row[COL_CATEGORIES]).strip() if row[COL_CATEGORIES] else ""
        raw_media = str(row[COL_MEDIA]).strip() if row[COL_MEDIA] else ""

        q_type = "basic" if structure == "PODSTAWOWY" else "specialist"
        media_name, media_type = resolve_media(raw_media if raw_media else None, media_dir_for_check)
        if raw_media and media_name is None:
            missing_media_count += 1

        # Build question object
        q_obj: dict = {
            "id": int(qnum) if qnum.isdigit() else qnum,
            "q": question_text,
            "type": q_type,
            "correct": correct,
            "media": media_name,
            "mediaType": media_type,
        }

        # Add ABC answers for specialist questions
        if q_type == "specialist":
            ans_a = str(row[COL_A]).strip() if row[COL_A] else ""
            ans_b = str(row[COL_B]).strip() if row[COL_B] else ""
            ans_c = str(row[COL_C]).strip() if row[COL_C] else ""
            q_obj["a"] = ans_a
            q_obj["b"] = ans_b
            q_obj["c"] = ans_c

        # Assign to each listed category
        for cat in raw_cats.split(","):
            cat = cat.strip()
            if cat in cat_questions:
                cat_questions[cat].append(q_obj)

    if missing_media_count:
        print(f"  WARNING: {missing_media_count} questions reference media files not found in source directory")

    # Write per-category JSON files
    print()
    meta_categories = []
    for cat in CATEGORIES:
        questions = cat_questions[cat]
        basic = [q for q in questions if q["type"] == "basic"]
        specialist = [q for q in questions if q["type"] == "specialist"]

        cat_file = out_dir / f"{cat}.json"
        with open(cat_file, "w", encoding="utf-8") as f:
            json.dump(
                {"category": cat, "questions": questions},
                f,
                ensure_ascii=False,
                indent=2,
            )

        meta_categories.append({
            "id": cat,
            "name": f"Kategoria {cat}",
            "questionCount": len(questions),
            "basicCount": len(basic),
            "specialistCount": len(specialist),
        })

        print(f"  {cat:>3}: {len(questions):>4} questions ({len(basic)} basic + {len(specialist)} specialist) → {cat_file.name}")

    # Write meta.json
    meta = {
        "categories": meta_categories,
        "exam": EXAM_RULES,
    }
    meta_file = out_dir / "meta.json"
    with open(meta_file, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    print(f"\n  meta.json written to {meta_file}")
    total = sum(c["questionCount"] for c in meta_categories)
    print(f"\n  TOTAL: {total} question-category assignments across {len(CATEGORIES)} categories")
    print("  Done!")


if __name__ == "__main__":
    main()
